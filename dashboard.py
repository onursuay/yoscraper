#!/usr/bin/env python3
"""
Google Maps Isletme Tarayici - Web Dashboard
"""
from datetime import datetime, timedelta

import os
import json
import threading
import logging
import warnings
from flask import Flask, render_template, request, jsonify, redirect, session, make_response, g
from flask_socketio import SocketIO

from scraper.maps_scraper import BusinessScraper
from scraper.email_extractor import EmailExtractor
from sheets.sheets_manager import SheetsManager
from sheets.leads_manager import LeadsManager
from utils.filters import extract_domain_from_url
from utils.domain_parser import domain_to_business_name

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("dashboard")

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.getenv("FLASK_SECRET_KEY", "ps-dev-secret-change-in-prod")
app.config["SESSION_PERMANENT"] = True
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

from auth_utils import (
    get_user_by_email, get_user_by_id, create_user,
    update_user, update_password, verify_password, get_initials,
)

# ─── Auth helpers ─────────────────────────────────────────────────────────────

_PROTECTED = ('/app', '/external', '/database', '/domains', '/sendmail',
               '/marketing', '/clickbot', '/account',
               '/faturalarim', '/abonelik', '/yardim', '/dashboard')


@app.before_request
def require_auth():
    path = request.path
    if not any(path == p or path.startswith(p + '/') for p in _PROTECTED):
        return
    if session.get('user_id'):
        return
    if request.cookies.get('auth') == '1':
        return  # legacy cookie grace period
    if request.content_type and 'json' in request.content_type:
        return jsonify({'error': 'Unauthorized'}), 401
    return redirect('/login')


@app.context_processor
def inject_user():
    if not hasattr(g, '_current_user'):
        uid = session.get('user_id')
        if uid and uid != '__admin__':
            user = get_user_by_id(uid)
            if user:
                user['initials'] = get_initials(user.get('full_name', '?'))
            g._current_user = user
        else:
            g._current_user = None
    return {'current_user': g._current_user}

# Click Bot entegrasyonu
from clickbot.routes import clickbot_bp, init_clickbot, register_socketio_handlers
app.register_blueprint(clickbot_bp)
init_clickbot(socketio)
register_socketio_handlers(socketio)

from marketing.scheduler import start_scheduler
from marketing.db import is_suppressed, add_suppression, remove_suppression
from marketing.unsub import generate_token, verify_token
from marketing.queue import build_footer
from marketing.routes import marketing_bp
app.register_blueprint(marketing_bp)

# Tarama durumu (global)
scan_state = {
    "running": False,
    "paused": False,
    "progress": 0,
    "total": 0,
    "status": "Hazır",
    "results": [],
    "logs": [],
}

# scan_state'i app.config'e kaydet (clickbot erişebilsin)
app.config["SCAN_STATE"] = scan_state

# Threading events for pause/stop
import threading as _scan_threading
_scan_pause_event = _scan_threading.Event()
_scan_stop_event = _scan_threading.Event()

# Sektor ikonlari (SVG path)
SECTOR_ICONS = {
    "Güzellik, Kişisel Bakım, Wellness": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20.84 4.61a5.5 5.5 0 0 0-7.78 0L12 5.67l-1.06-1.06a5.5 5.5 0 0 0-7.78 7.78l1.06 1.06L12 21.23l7.78-7.78 1.06-1.06a5.5 5.5 0 0 0 0-7.78z"/></svg>',
    "Yeme-İçme": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M18 8h1a4 4 0 0 1 0 8h-1"/><path d="M2 8h16v9a4 4 0 0 1-4 4H6a4 4 0 0 1-4-4V8z"/><line x1="6" y1="1" x2="6" y2="4"/><line x1="10" y1="1" x2="10" y2="4"/><line x1="14" y1="1" x2="14" y2="4"/></svg>',
    "Konaklama Ve Turizm": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 21h18"/><path d="M5 21V7l8-4v18"/><path d="M19 21V11l-6-4"/><path d="M9 9h.01"/><path d="M9 13h.01"/><path d="M9 17h.01"/></svg>',
    "Otomotiv": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M5 17h14M5 17a2 2 0 0 1-2-2V7h18v8a2 2 0 0 1-2 2M5 17l-1 3h16l-1-3"/><circle cx="7.5" cy="14.5" r="1.5"/><circle cx="16.5" cy="14.5" r="1.5"/></svg>',
    "İnşaat, Emlak, Mimarlık": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>',
    "Kurumsal Hizmetler": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="2" y="7" width="20" height="14" rx="2" ry="2"/><path d="M16 21V5a2 2 0 0 0-2-2h-4a2 2 0 0 0-2 2v16"/></svg>',
    "Eğitim": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M2 3h6a4 4 0 0 1 4 4v14a3 3 0 0 0-3-3H2z"/><path d="M22 3h-6a4 4 0 0 0-4 4v14a3 3 0 0 1 3-3h7z"/></svg>',
    "Finans, Sigorta, Hukuk, Danışmanlık": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>',
    "Teknoloji Ve Dijital": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="16 18 22 12 16 6"/><polyline points="8 6 2 12 8 18"/></svg>',
    "Medya, Organizasyon, Spor": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="23 7 16 12 23 17 23 7"/><rect x="1" y="5" width="15" height="14" rx="2" ry="2"/></svg>',
    "Üretim Ve Sanayi": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2v4"/><path d="M6.8 15a6 6 0 1 1 10.4 0"/><line x1="2" y1="22" x2="22" y2="22"/><path d="M7 22v-4"/><path d="M17 22v-4"/></svg>',
    "Enerji Ve Çevre": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="5"/><line x1="12" y1="1" x2="12" y2="3"/><line x1="12" y1="21" x2="12" y2="23"/><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"/><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"/><line x1="1" y1="12" x2="3" y2="12"/><line x1="21" y1="12" x2="23" y2="12"/><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"/><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"/></svg>',
    "Lojistik": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="1" y="3" width="15" height="13"/><polygon points="16 8 20 8 23 11 23 16 16 16 16 8"/><circle cx="5.5" cy="18.5" r="2.5"/><circle cx="18.5" cy="18.5" r="2.5"/></svg>',
    "Sağlık Ve Medikal": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 12h-4l-3 9L9 3l-3 9H2"/></svg>',
    "Perakende Ve E-Ticaret": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="9" cy="21" r="1"/><circle cx="20" cy="21" r="1"/><path d="M1 1h4l2.68 13.39a2 2 0 0 0 2 1.61h9.72a2 2 0 0 0 2-1.61L23 6H6"/></svg>',
}

# Sektorler - Sadece kurumsal web sitesi/e-posta olma ihtimali yuksek olanlar
SECTORS = {
    "Güzellik, Kişisel Bakım, Wellness": [
        "Güzellik Merkezi", "Cilt Bakımı Merkezi", "Lazer Epilasyon",
        "Zayıflama Merkezi", "SPA Merkezi", "Hamam / Termal SPA",
        "Nail Studio (Manikür-Pedikür)", "Kalıcı Makyaj / Microblading",
        "Kozmetik Mağazası", "Kişisel Bakım Ürünleri E-Ticaret",
    ],
    "Yeme-İçme": [
        "Restoran", "Kafe Zinciri", "Pastane", "Fast Food",
        "Dönerci", "Kebapçı", "Burgerci", "Pizza",
        "Balık Restoranı", "Steakhouse", "Kahve Zinciri / Coffee Shop",
        "Catering (Kurumsal Yemek)", "Yemek Üretim Tesisi",
        "Gıda Toptancısı",
    ],
    "Konaklama Ve Turizm": [
        "Otel", "Butik Otel", "Apart Otel", "Tatil Köyü",
        "Termal Otel", "Seyahat Acentası", "Tur Operatörü",
        "Transfer Firması", "Vize Danışmanlığı",
    ],
    "Otomotiv": [
        "Oto Servis", "Oto Yıkama", "Oto Detailing / Seramik Kaplama",
        "Oto Kiralama (Rent A Car)", "Filo Kiralama",
        "Galeri / İkinci El Araç Satışı", "Sıfır Araç Bayii",
        "Yedek Parça Satışı", "Motosiklet Bayii",
    ],
    "İnşaat, Emlak, Mimarlık": [
        "Müteahhitlik / İnşaat Firması", "Emlak Ofisi",
        "Mimarlık Ofisi", "İç Mimarlık", "İnşaat Mühendisliği",
        "Prefabrik Yapı", "İnşaat Malzemeleri Bayii",
        "Kapı / Pencere (PVC, Alüminyum)", "Cam Balkon",
        "Asansör Firması", "Peyzaj Firması",
        "Çatı / İzolasyon", "Mantolama",
    ],
    "Kurumsal Hizmetler": [
        "Temizlik Şirketi", "Dezenfeksiyon Hizmetleri", "İlaçlama (Haşere)",
        "Kamera / Alarm Sistemleri", "Yangın Sistemleri",
        "Nakliyat / Evden Eve Taşıma", "Ofis Taşıma",
    ],
    "Eğitim": [
        "Anaokulu / Kreş", "Özel Okul", "Kolej",
        "YKS/LGS Kursu", "Dil Kursu", "Sürücü Kursu",
        "Yazılım Kursu / Bootcamp", "Online Eğitim Platformu",
        "Kurumsal Eğitim / İK Eğitimleri",
    ],
    "Finans, Sigorta, Hukuk, Danışmanlık": [
        "Muhasebe Ofisi / Mali Müşavir", "Sigorta Acentası",
        "Finansal Danışmanlık", "Yatırım Danışmanlığı",
        "Hukuk Bürosu", "Patent / Marka Tescil",
        "Yönetim Danışmanlığı", "İnsan Kaynakları Danışmanlığı",
    ],
    "Teknoloji Ve Dijital": [
        "Dijital Pazarlama Ajansı", "SEO Ajansı", "Sosyal Medya Ajansı",
        "Kreatif Ajans / Tasarım Stüdyosu", "Yazılım Şirketi",
        "Web Tasarım / UI-UX Stüdyosu", "Mobil Uygulama Geliştirme",
        "IT Destek / Sistem Entegratörü", "Siber Güvenlik Hizmetleri",
        "Bulut Hizmetleri / Hosting", "Çağrı Merkezi",
        "ERP Danışmanlık", "E-Ticaret Altyapı Sağlayıcı",
    ],
    "Medya, Organizasyon, Spor": [
        "Prodüksiyon Şirketi", "Fotoğraf Stüdyosu",
        "Düğün Organizasyon", "Etkinlik Organizasyon",
        "Fuar Stand Tasarım", "Spor Salonu",
        "Pilates / Yoga Stüdyosu",
    ],
    "Üretim Ve Sanayi": [
        "Gıda Üretimi", "Tekstil Üretimi", "Hazır Giyim Üretimi",
        "Mobilya Üretimi", "Metal İşleme", "Makine İmalat",
        "Otomotiv Yan Sanayi", "Plastik Enjeksiyon", "Ambalaj Üretimi",
        "Kimya Üretimi", "Elektrik-Elektronik Üretim",
        "Matbaa / Baskı", "Reklam Tabela İmalatı", "Medikal Üretim",
    ],
    "Enerji Ve Çevre": [
        "Güneş Enerjisi Kurulum (GES)", "Elektrik Taahhüt",
        "Jeneratör Satış/Servis", "Geri Dönüşüm",
        "Su Arıtma Sistemleri", "Çevre Danışmanlık",
    ],
    "Lojistik": [
        "Yurtiçi Nakliye", "Uluslararası Taşımacılık",
        "Gümrük Müşavirliği", "Kargo Firması",
        "Soğuk Zincir Lojistik", "Depo / Antrepo",
    ],
    "Sağlık Ve Medikal": [
        "Özel Hastane", "Tıp Merkezi / Poliklinik",
        "Saç Ekimi Merkezi", "Estetik Cerrahi",
        "Medikal Estetik (Botoks, Dolgu)", "Diş Kliniği",
        "Göz Hastalıkları / Lazer Göz", "Fizik Tedavi",
        "Psikolog / Terapi Merkezi", "Diyetisyen",
        "Tıbbi Laboratuvar", "Medikal Cihaz Satışı",
        "Veteriner Kliniği",
    ],
    "Perakende Ve E-Ticaret": [
        "Giyim Mağazası", "Ayakkabı Mağazası", "Optik / Gözlük",
        "Mobilya Mağazası", "Elektronik Perakende", "Beyaz Eşya Bayii",
        "Yapı Market", "E-Ticaret (Genel)", "E-Ticaret (D2C Marka)",
    ],
}

CITIES = [
    "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya",
    "Adana", "Konya", "Gaziantep", "Mersin", "Kayseri",
    "Diyarbakır", "Eskişehir", "Samsun", "Denizli", "Şanlıurfa",
    "Trabzon", "Malatya", "Erzurum", "Van", "Batman",
    "Elazığ", "Kocaeli", "Sakarya", "Tekirdağ", "Muğla",
    "Balıkesir", "Manisa", "Aydın", "Hatay", "Kahramanmaraş",
    "Afyonkarahisar", "Sivas", "Tokat", "Yozgat", "Aksaray",
    "Kırşehir", "Nevşehir", "Niğde", "Karaman", "Konya",
    "Düzce", "Bolu", "Zonguldak", "Bartın", "Kastamonu",
    "Çankırı", "Çorum", "Amasya", "Giresun", "Ordu",
    "Rize", "Artvin", "Gümüşhane", "Bayburt", "Erzincan",
    "Tunceli", "Bingöl", "Muş", "Ağrı", "Iğdır",
    "Kars", "Ardahan", "Siirt", "Şırnak", "Hakkari",
    "Bitlis", "Mardin", "Adıyaman", "Osmaniye", "Kilis",
    "Gaziantep", "Isparta", "Burdur", "Uşak", "Kütahya",
    "Bilecik", "Edirne", "Kırklareli", "Çanakkale", "Yalova",
    "Sinop", "Karabük",
]
# Tekrarlari kaldir ve sirala
CITIES = sorted(set(CITIES), key=lambda x: x.replace("İ", "I").replace("Ş", "S"))


def add_log(message: str):
    scan_state["logs"].append(message)
    if len(scan_state["logs"]) > 100:
        scan_state["logs"] = scan_state["logs"][-50:]


def _auto_trigger_scanner_leads(businesses: list[dict]):
    """Scanner'dan gelen yeni leadleri trigger_type='lead_created' kampanyalarına gönder."""
    try:
        from marketing.campaigns import auto_trigger_leads
        leads = [
            {
                "email":  b.get("email", ""),
                "name":   b.get("name", ""),
                "sector": b.get("sector", ""),
                "city":   b.get("city", ""),
                "domain": b.get("domain", ""),
                "website": b.get("website", ""),
                "first_name": "",
                "last_name":  "",
            }
            for b in businesses if b.get("email")
        ]
        n = auto_trigger_leads(leads, "scanner")
        if n:
            logger.info(f"auto_trigger: {n} scanner lead için kampanya tetiklendi")
    except Exception as e:
        logger.debug(f"auto_trigger_scanner error: {e}")


def run_scan(sector: str, city: str, min_results: int):
    """Arka planda tarama calistir."""
    scan_state["running"] = True
    scan_state["paused"] = False
    scan_state["progress"] = 0
    scan_state["total"] = 0
    scan_state["status"] = "Google Sheets'e bağlanılıyor..."
    scan_state["results"] = []
    scan_state["logs"] = []
    _scan_stop_event.clear()
    _scan_pause_event.clear()

    try:
        # 1. Sheets baglantisi
        add_log("Google Sheets'e bağlanılıyor...")
        sheets = SheetsManager()
        add_log(f"Bağlantı başarılı. Mevcut {len(sheets.existing_domains)} domain var.")

        # 2. Google Maps taramasi
        scan_state["status"] = f"Google Maps'te aranıyor: {sector} {city}..."
        add_log(f"Google Maps'te aranıyor: '{sector} {city}'")

        scraper = BusinessScraper()
        try:
            scraper.start_browser()
            # E-posta bulamayacaklari da hesaba katarak 3 kati isletme topla
            scraper_target = min_results * 3
            raw_businesses = scraper.search_businesses(sector, city, scraper_target)
        finally:
            scraper.close_browser()

        if not raw_businesses:
            scan_state["status"] = "Hiçbir işletme bulunamadı!"
            add_log("Hiçbir işletme bulunamadı.")
            scan_state["running"] = False
            return

        scan_state["total"] = len(raw_businesses)
        add_log(f"{len(raw_businesses)} web siteli işletme bulundu, e-postalar çıkarılıyor...")

        # 3. E-posta cikarma - min_results kurumsal e-posta bulana kadar devam et
        scan_state["status"] = "E-postalar çıkarılıyor..."
        extractor = EmailExtractor()
        valid_businesses = []

        for i, biz in enumerate(raw_businesses, 1):
            # Stop kontrolu
            if _scan_stop_event.is_set():
                add_log("Tarama kullanıcı tarafından durduruldu.")
                break

            # Pause kontrolu - pause kaldirilana kadar bekle
            while _scan_pause_event.is_set():
                if _scan_stop_event.is_set():
                    break
                import time
                time.sleep(0.5)

            if _scan_stop_event.is_set():
                add_log("Tarama kullanıcı tarafından durduruldu.")
                break

            # Yeterli sonuca ulastiysa dur
            if len(valid_businesses) >= min_results:
                add_log(f"Hedef {min_results} kurumsal e-postaya ulaşıldı!")
                break

            scan_state["progress"] = i
            website = biz.get("website", "")
            if not website:
                continue

            domain = extract_domain_from_url(website)

            if sheets.is_duplicate(domain):
                add_log(f"[{i}/{scan_state['total']}] MÜKERRER - {domain}")
                continue

            scan_state["status"] = f"[{i}/{scan_state['total']}] {domain} taranıyor... ({len(valid_businesses)}/{min_results} e-posta bulundu)"
            add_log(f"[{i}/{scan_state['total']}] {domain} taranıyor...")

            emails = extractor.extract_emails_from_url(website)
            if not emails:
                add_log(f"  → Kurumsal e-posta bulunamadı.")
                continue

            # Firma adini belirle (3 asamali):
            # 1. Google Places API'den gelen ad (en guvenilir)
            maps_name = biz.get("maps_name", "").strip()
            # Uzun Google isimlerini temizle (| ve - ile ayrilmis kisimlari kes)
            if maps_name:
                for sep in [" | ", " - ", " – "]:
                    if sep in maps_name:
                        maps_name = maps_name.split(sep)[0].strip()
                        break
                # Cok uzunsa kirp
                if len(maps_name) > 50:
                    maps_name = maps_name[:50].rsplit(" ", 1)[0]

            # 2. Web sitesinin <title> etiketinden
            site_title = extractor.extract_site_title(website)

            # 3. Domain'den tahmin (son care)
            domain_name = domain_to_business_name(website)

            # En iyi adi sec: Google Places > Site Title > Domain
            business_name = maps_name or site_title or domain_name
            email = emails[0]

            # Sosyal medya linklerini cikar
            social = extractor.extract_social_links(website)
            today = datetime.now().strftime("%Y-%m-%d")

            result = {
                "date": today,
                "sector": sector,
                "name": business_name,
                "phone": biz.get("phone", ""),
                "email": email,
                "domain": domain,
                "website": website,
                "instagram": social.get("instagram", ""),
                "facebook": social.get("facebook", ""),
                "linkedin": social.get("linkedin", ""),
            }
            valid_businesses.append(result)
            scan_state["results"].append(result)
            add_log(f"  → BULUNDU: {business_name} - {email}")

        # 4. Sheets'e kaydet
        if valid_businesses:
            scan_state["status"] = "Google Sheets'e kaydediliyor..."
            add_log("Google Sheets'e kaydediliyor...")
            added = sheets.append_businesses(valid_businesses)
            add_log(f"{added} yeni işletme Google Sheets'e eklendi.")
            scan_state["status"] = f"Tamamlandı! {len(valid_businesses)} işletme bulundu, {added} yeni kayıt eklendi."

            # Auto-enroll: lead_created trigger olan sequence kampanyalarına enroll et
            _auto_trigger_scanner_leads(valid_businesses)
        elif len(scan_state["results"]) > 0:
            # Sonuclar bulundu ama hepsi zaten kayitli
            scan_state["status"] = f"Tamamlandı! {len(scan_state['results'])} işletme bulundu (hepsi zaten kayıtlı)."
            add_log(f"{len(scan_state['results'])} işletme bulundu ama hepsi zaten Google Sheets'te kayıtlı.")
        else:
            scan_state["status"] = "Tamamlandı - Bu arama için kurumsal e-posta bulunamadı."
            add_log("Kurumsal e-posta bulunamadı.")

        # Playwright tarayicisini kapat
        extractor.close()

    except Exception as e:
        scan_state["status"] = f"Hata: {str(e)}"
        add_log(f"HATA: {str(e)}")
        logger.exception("Tarama hatası")

    finally:
        scan_state["running"] = False
        scan_state["paused"] = False
        _scan_stop_event.clear()
        _scan_pause_event.clear()


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get('user_id'):
        return redirect('/app')
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        email = data.get("email", "").strip().lower()
        password = data.get("password", "")

        if not email or not password:
            return jsonify({"error": "E-posta ve şifre gereklidir."}), 400

        user = get_user_by_email(email)
        if user and verify_password(password, user.get("password_hash", "")):
            session['user_id'] = user['id']
            session.permanent = True
            resp = make_response(jsonify({"success": True}))
            resp.set_cookie("auth", "", expires=0)
            return resp

        # APP_PASSWORD fallback — admin login without user record
        app_password = os.getenv("APP_PASSWORD", "")
        if app_password and password == app_password:
            session['user_id'] = '__admin__'
            session.permanent = True
            return jsonify({"success": True})

        return jsonify({"error": "E-posta veya şifre hatalı."}), 401
    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if session.get('user_id'):
        return redirect('/app')
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        name     = data.get("name", "").strip()
        email    = data.get("email", "").strip()
        company  = data.get("company", "").strip()
        phone    = data.get("phone", "").strip()
        password = data.get("password", "")

        if not name or not email or not password:
            return jsonify({"error": "Ad soyad, e-posta ve şifre gereklidir."}), 400
        if len(password) < 8:
            return jsonify({"error": "Şifre en az 8 karakter olmalıdır."}), 400

        result = create_user(name, email, company, phone, password)
        if "error" in result:
            return jsonify(result), 400

        session['user_id'] = result['id']
        session.permanent = True
        resp = make_response(jsonify({"success": True}))
        resp.set_cookie("auth", "", expires=0)
        return resp
    return render_template("register.html")


@app.route("/logout")
def logout():
    session.clear()
    resp = make_response(redirect('/'))
    resp.set_cookie("auth", "", expires=0)
    return resp


@app.route("/account", methods=["GET", "POST"])
@app.route("/hesabim", methods=["GET", "POST"])
def account():
    uid = session.get('user_id')
    if not uid or uid == '__admin__':
        return redirect('/login')
    user = get_user_by_id(uid)
    if not user:
        session.clear()
        return redirect('/login')

    if request.method == "POST":
        data   = request.get_json(silent=True) or {}
        action = data.get("action", "update_profile")

        if action == "update_profile":
            result = update_user(uid, {
                "full_name": data.get("full_name", user["full_name"]).strip(),
                "company":   data.get("company",   user.get("company", "")).strip(),
                "phone":     data.get("phone",     user.get("phone",   "")).strip(),
            })
            return jsonify({"success": True, "user": result})

        if action == "change_password":
            cur_pw = data.get("current_password", "")
            new_pw = data.get("new_password", "")
            if not verify_password(cur_pw, user["password_hash"]):
                return jsonify({"error": "Mevcut şifre hatalı."}), 400
            if len(new_pw) < 8:
                return jsonify({"error": "Yeni şifre en az 8 karakter olmalıdır."}), 400
            update_password(uid, new_pw)
            return jsonify({"success": True})

        return jsonify({"error": "Bilinmeyen işlem."}), 400

    user['initials'] = get_initials(user.get('full_name', '?'))
    return render_template("account.html", user=user)


@app.route("/faturalarim")
def faturalarim():
    return render_template("faturalarim.html")


@app.route("/abonelik")
def abonelik():
    return render_template("abonelik.html")


@app.route("/yardim")
def yardim():
    return render_template("yardim.html")


@app.route("/dashboard")
@app.route("/genel-bakis")
def app_dashboard():
    return render_template("app_dashboard.html")


@app.route("/app")
@app.route("/tarayici")
def index():
    return render_template("index.html", sectors=SECTORS, cities=CITIES, sector_icons=SECTOR_ICONS)


@app.route("/en/privacy-policy")
def privacy_policy():
    return render_template("privacy_policy.html")


@app.route("/en/terms-of-service")
def terms_of_service():
    return render_template("terms_of_service.html")


@app.route("/robots.txt")
def robots_txt():
    content = (
        "User-agent: *\n"
        "Allow: /\n"
        "Allow: /en/privacy-policy\n"
        "Allow: /en/terms-of-service\n\n"
        "Sitemap: https://scraper.yodijital.com/sitemap.xml\n"
    )
    return content, 200, {"Content-Type": "text/plain; charset=utf-8"}


@app.route("/sitemap.xml")
def sitemap_xml():
    content = '''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url>
    <loc>https://scraper.yodijital.com/</loc>
    <changefreq>weekly</changefreq>
    <priority>1.0</priority>
  </url>
  <url>
    <loc>https://scraper.yodijital.com/en/privacy-policy</loc>
    <changefreq>monthly</changefreq>
    <priority>0.5</priority>
  </url>
  <url>
    <loc>https://scraper.yodijital.com/en/terms-of-service</loc>
    <changefreq>monthly</changefreq>
    <priority>0.5</priority>
  </url>
</urlset>'''
    return content, 200, {"Content-Type": "application/xml; charset=utf-8"}


@app.route("/api/reset", methods=["POST"])
def reset_state():
    """Tarama durumunu sifirla."""
    scan_state["running"] = False
    scan_state["paused"] = False
    scan_state["progress"] = 0
    scan_state["total"] = 0
    scan_state["status"] = "Hazır"
    scan_state["results"] = []
    scan_state["logs"] = []
    _scan_stop_event.clear()
    _scan_pause_event.clear()
    return jsonify({"message": "Sıfırlandı"})


@app.route("/api/scan/pause", methods=["POST"])
def pause_scan():
    """Taramayı duraklat."""
    if scan_state["running"]:
        _scan_pause_event.set()
        scan_state["paused"] = True
        scan_state["status"] = "Duraklatıldı"
        scan_state["logs"].append("⏸ Tarama duraklatıldı")
    return jsonify({"status": "paused"})


@app.route("/api/scan/resume", methods=["POST"])
def resume_scan():
    """Taramaya devam et."""
    if scan_state["running"] and scan_state["paused"]:
        _scan_pause_event.clear()
        scan_state["paused"] = False
        scan_state["status"] = "Devam ediyor..."
        scan_state["logs"].append("▶ Tarama devam ediyor")
    return jsonify({"status": "resumed"})


@app.route("/api/scan/stop", methods=["POST"])
def stop_scan():
    """Taramayı durdur."""
    if scan_state["running"]:
        _scan_stop_event.set()
        _scan_pause_event.clear()  # Pause'da bekliyorsa çık
        scan_state["paused"] = False
        scan_state["status"] = "Durduruldu"
        scan_state["logs"].append("⏹ Tarama durduruldu")
    return jsonify({"status": "stopped"})


@app.route("/api/scan", methods=["POST"])
def start_scan():
    if scan_state["running"]:
        return jsonify({"error": "Tarama zaten devam ediyor!"}), 400

    # Onceki state'i sifirla
    scan_state["progress"] = 0
    scan_state["total"] = 0
    scan_state["status"] = "Başlatılıyor..."
    scan_state["results"] = []
    scan_state["logs"] = []

    data = request.json
    sector = data.get("sector", "")
    city = data.get("city", "")
    min_results = data.get("min_results", 20)

    if not sector or not city:
        return jsonify({"error": "Sektör ve şehir seçimi zorunludur!"}), 400

    thread = threading.Thread(
        target=run_scan,
        args=(sector, city, int(min_results)),
        daemon=True,
    )
    thread.start()

    return jsonify({"message": "Tarama başlatıldı!"})


@app.route("/api/scan/delete", methods=["POST"])
def delete_scan_results():
    """Secilen tarama sonuclarini sil (hem bellekten hem Sheets'ten)."""
    data = request.json
    indices = sorted(data.get("indices", []), reverse=True)
    if not indices:
        return jsonify({"error": "Silinecek kayıt seçilmedi"}), 400

    deleted = 0
    try:
        sheets = SheetsManager()
        for idx in indices:
            if idx < len(scan_state["results"]):
                result = scan_state["results"][idx]
                domain = result.get("domain", "")

                # Sheets'ten sil - domain'e gore bul
                try:
                    domain_col = sheets.worksheet.col_values(6)  # F = Domain
                    for row_idx, val in enumerate(domain_col):
                        if val.lower().strip() == domain.lower().strip():
                            sheets.worksheet.delete_rows(row_idx + 1)
                            break
                except Exception:
                    pass

                deleted += 1

        # Bellekten sil
        for idx in indices:
            if idx < len(scan_state["results"]):
                scan_state["results"].pop(idx)

        return jsonify({"message": f"{deleted} kayıt silindi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/status")
def get_status():
    return jsonify({
        "running": scan_state["running"],
        "paused": scan_state.get("paused", False),
        "progress": scan_state["progress"],
        "total": scan_state["total"],
        "status": scan_state["status"],
        "results": scan_state["results"],
        "logs": scan_state["logs"][-20:],
    })


@app.route("/api/sectors")
def get_sectors():
    return jsonify(SECTORS)


@app.route("/api/domains")
def get_domains():
    """Google Sheets'teki tum kayitli domainleri getir."""
    try:
        sheets = SheetsManager()
        records = sheets.worksheet.get_all_values()
        # Basligi atla
        data = []
        for i, row in enumerate(records[1:], 1):
            if len(row) >= 4 and row[0]:
                data.append({
                    "row": i + 1,  # Sheets'teki satir numarasi
                    "name": row[0],
                    "email": row[1] if len(row) > 1 else "",
                    "phone": row[2] if len(row) > 2 else "",
                    "domain": row[3] if len(row) > 3 else "",
                })
        return jsonify({"domains": data, "total": len(data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/domains/delete", methods=["POST"])
def delete_domains():
    """Secilen satirlari Google Sheets'ten tek batch istegiyle sil."""
    try:
        data = request.json
        rows_to_delete = sorted(data.get("rows", []), reverse=True)  # Sondan basa sil
        if not rows_to_delete:
            return jsonify({"error": "Silinecek satir secilmedi"}), 400

        sheets = SheetsManager()
        spreadsheet_id = sheets.worksheet.spreadsheet.id
        sheet_id = sheets.worksheet.id

        # Tek bir batchUpdate istegiyle tum satirlari sil (API quota asmaz)
        requests_body = []
        for row_num in rows_to_delete:
            requests_body.append({
                "deleteDimension": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "ROWS",
                        "startIndex": row_num - 1,  # 0-indexed
                        "endIndex": row_num          # exclusive
                    }
                }
            })

        from googleapiclient.discovery import build
        from config import get_google_credentials
        creds = get_google_credentials([
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ])
        service = build("sheets", "v4", credentials=creds)
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests_body}
        ).execute()

        return jsonify({"message": f"{len(rows_to_delete)} kayıt silindi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/domains")
def domains_page():
    return render_template("domains.html")


@app.route("/external")
@app.route("/ice-aktarma")
def external_page():
    lead_url = app.config.get("LEAD_SHEET_URL", "")
    lead_sheet = app.config.get("LEAD_SHEET_NAME", "")
    return render_template("external.html", has_connection=(bool(lead_url and lead_sheet)), lead_sheet_name=lead_sheet)


@app.route("/database")
@app.route("/veritabani")
def database_page():
    return render_template("database.html")


@app.route("/sendmail")
def sendmail_page():
    return render_template("sendmail.html")


# --- Unsubscribe ---
@app.route("/unsubscribe")
def unsubscribe_page():
    token = request.args.get("token", "")
    email_raw = request.args.get("email", "")  # eski link backward compat

    email = ""
    error = False

    if token:
        email = verify_token(token) or ""
        if not email:
            error = True
    elif email_raw:
        email = email_raw.lower().strip()

    if email and not error:
        add_suppression(email, reason="unsubscribed", source="user")

    return render_template("unsubscribe.html", email=email, error=error)


@app.route("/api/unsubscribed", methods=["GET"])
def get_unsubscribed():
    from marketing.db import sb_select
    try:
        rows = sb_select("email_suppressions", {
            "reason": "eq.unsubscribed",
            "select": "email,created_at",
            "order": "created_at.desc",
        })
        return jsonify({"emails": [r["email"] for r in rows]})
    except Exception as e:
        return jsonify({"emails": [], "error": str(e)})


@app.route("/api/unsubscribed", methods=["DELETE"])
def remove_unsubscribed():
    email = (request.json or {}).get("email", "")
    if email:
        remove_suppression(email)
    return jsonify({"ok": True})


# --- Send Mail API ---
_mail_history = []  # [{id, to, name, subject, status, timestamp, error}]
_mail_sending = False


@app.route("/api/sendmail/send", methods=["POST"])
def api_sendmail_send():
    """Secilen firmalara mail gonder."""
    global _mail_sending
    if _mail_sending:
        return jsonify({"error": "Gönderim devam ediyor, lütfen bekleyin."}), 409

    data = request.json or {}
    recipients = data.get("recipients", [])  # [{email, name}]
    subject = data.get("subject", "")
    html_body = data.get("html_body", "")

    if not recipients:
        return jsonify({"error": "Alıcı seçilmedi."}), 400
    if not subject:
        return jsonify({"error": "Konu girilmedi."}), 400
    if not html_body:
        return jsonify({"error": "Mail içeriği boş."}), 400

    import os
    api_key = os.getenv("RESEND_API_KEY", "")
    if not api_key:
        return jsonify({"error": "RESEND_API_KEY ayarlanmamış. .env dosyasını kontrol edin."}), 400

    from_name = os.getenv("FROM_NAME", "YO Dijital")
    from_email = os.getenv("FROM_EMAIL", "info@yodijital.com")

    base_url = os.getenv("APP_BASE_URL", "https://scraper.yodijital.com")

    def send_thread():
        global _mail_sending
        _mail_sending = True
        import requests as _req
        import time

        for r in recipients:
            email = r.get("email", "")
            name = r.get("name", "Yetkili")
            if not email:
                continue

            # Suppression kontrolü (Supabase email_suppressions)
            if is_suppressed(email):
                entry = {
                    "id": len(_mail_history) + 1,
                    "to": email,
                    "name": name,
                    "subject": subject,
                    "status": "failed",
                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                    "error": "suppressed",
                }
                _mail_history.append(entry)
                socketio.emit("mail_update", entry)
                continue

            sector = r.get("sector", "")
            city = r.get("city", "")
            body = html_body.replace("{firma_adi}", name).replace("{from_name}", from_name).replace("{email}", email).replace("{sektor}", sector).replace("{sehir}", city)

            # Footer + List-Unsubscribe
            footer_html, unsub_url = build_footer(base_url, email)
            body = body + footer_html

            entry = {
                "id": len(_mail_history) + 1,
                "to": email,
                "name": name,
                "subject": subject,
                "status": "sending",
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "error": "",
            }

            try:
                resp = _req.post(
                    "https://api.resend.com/emails",
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "from": f"{from_name} <{from_email}>",
                        "to": [email],
                        "subject": subject,
                        "html": body,
                        "headers": {
                            "List-Unsubscribe": f"<{unsub_url}>",
                            "List-Unsubscribe-Post": "List-Unsubscribe=One-Click",
                        },
                    },
                    timeout=15,
                )
                if resp.status_code == 200:
                    entry["status"] = "sent"
                else:
                    entry["status"] = "failed"
                    entry["error"] = resp.json().get("message", resp.text)[:100]
            except Exception as e:
                entry["status"] = "failed"
                entry["error"] = str(e)[:100]

            _mail_history.append(entry)
            socketio.emit("mail_update", entry)
            time.sleep(1)

        _mail_sending = False
        socketio.emit("mail_done", {"total": len(recipients)})

    threading.Thread(target=send_thread, daemon=True).start()
    return jsonify({"ok": True, "count": len(recipients)})


@app.route("/api/sendmail/template")
def api_sendmail_template():
    """Email template dosyasini oku."""
    import os
    template_file = os.getenv("EMAIL_TEMPLATE_FILE", "email_template.html")
    template_path = os.path.join(os.path.dirname(__file__), template_file)
    try:
        with open(template_path, "r", encoding="utf-8") as f:
            return jsonify({"html": f.read()})
    except Exception:
        return jsonify({"html": ""})


@app.route("/api/sendmail/history")
def api_sendmail_history():
    """Mail gonderim gecmisini getir."""
    return jsonify(_mail_history)


@app.route("/api/sendmail/status")
def api_sendmail_status():
    """Gonderim durumu."""
    sent = sum(1 for m in _mail_history if m["status"] == "sent")
    failed = sum(1 for m in _mail_history if m["status"] == "failed")
    return jsonify({"sending": _mail_sending, "total": len(_mail_history), "sent": sent, "failed": failed})


@app.route("/api/database/connection-status")
def connection_status():
    """Mevcut harici baglanti durumunu dondur."""
    lead_url = app.config.get("LEAD_SHEET_URL", "")
    lead_sheet = app.config.get("LEAD_SHEET_NAME", "")

    if not lead_url or not lead_sheet:
        return jsonify({"connected": False})

    try:
        lead_data = _fetch_lead_sheet(None, lead_url, lead_sheet)
        # Spreadsheet basligini al
        import gspread
        from config import get_google_credentials
        creds = get_google_credentials(["https://www.googleapis.com/auth/spreadsheets.readonly", "https://www.googleapis.com/auth/drive.readonly"])
        client = gspread.authorize(creds)
        sp = client.open_by_url(lead_url)

        return jsonify({
            "connected": True,
            "sheet_name": lead_sheet,
            "spreadsheet_title": sp.title,
            "url": lead_url,
            "count": len(lead_data),
        })
    except Exception as e:
        return jsonify({"connected": True, "sheet_name": lead_sheet, "spreadsheet_title": "?", "url": lead_url, "count": 0, "error": str(e)})


@app.route("/api/database/disconnect", methods=["POST"])
def disconnect_lead():
    """Harici baglantiyi kes."""
    app.config["LEAD_SHEET_URL"] = ""
    app.config["LEAD_SHEET_NAME"] = ""

    # .env'den kaldir
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        if os.path.exists(env_path):
            with open(env_path, "r") as f:
                lines = f.readlines()
            lines = [l for l in lines if not l.startswith("LEAD_SHEET_URL=") and not l.startswith("LEAD_SHEET_NAME=")]
            with open(env_path, "w") as f:
                f.writelines(lines)
    except Exception:
        pass

    return jsonify({"message": "Bağlantı kesildi"})


@app.route("/api/database/upload-excel", methods=["POST"])
def upload_excel():
    """Excel dosyasi yukle ve Database'e (Google Sheets) ekle."""
    if "file" not in request.files:
        return jsonify({"error": "Dosya secilmedi."}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Dosya secilmedi."}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        return jsonify({"error": "Desteklenen formatlar: .xlsx, .xls, .csv"}), 400

    try:
        import io
        from datetime import datetime as _dt

        rows_data = []

        if ext == "csv":
            import csv
            content = file.read().decode("utf-8-sig")
            reader = csv.reader(io.StringIO(content))
            all_rows = list(reader)
            if not all_rows:
                return jsonify({"error": "Dosya bos."}), 400
            headers = [h.strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return jsonify({"error": "Dosya bos."}), 400
            headers = [str(h or "").strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]

        # Kolon eslemesi
        col_map = _map_excel_columns(headers)

        if "email" not in col_map and "name" not in col_map:
            return jsonify({"error": f"Tanınan sütun bulunamadı. Mevcut başlıklar: {', '.join(headers)}"}), 400

        today = _dt.now().strftime("%Y-%m-%d")

        for row in data_rows:
            row = list(row) if not isinstance(row, list) else row
            if not any(str(c or "").strip() for c in row):
                continue

            def _get(field):
                idx = col_map.get(field)
                if idx is not None and idx < len(row):
                    return str(row[idx] or "").strip()
                return ""

            email = _get("email")
            name = _get("name")
            if not email and not name:
                continue

            rows_data.append({
                "name": name or "",
                "email": email or "",
                "phone": _get("phone") or "",
                "domain": _get("domain") or "",
                "sector": _get("sector") or "Excel Import",
                "website": _get("website") or "",
                "instagram": _get("instagram") or "",
                "facebook": _get("facebook") or "",
                "linkedin": _get("linkedin") or "",
                "date": _get("date") or today,
            })

        if not rows_data:
            return jsonify({"error": "Gecerli veri bulunamadi."}), 400

        # Google Sheets'e ekle
        sheets = SheetsManager()
        added = sheets.append_businesses(rows_data)

        return jsonify({
            "message": f"{added} kayıt Database'e eklendi ({len(rows_data)} satır okundu, {len(rows_data) - added} mükerrer atlandı).",
            "added": added,
            "total_read": len(rows_data),
            "duplicates": len(rows_data) - added,
        })

    except Exception as e:
        logger.error(f"Excel upload hatasi: {e}")
        return jsonify({"error": f"Dosya isleme hatasi: {str(e)}"}), 500


def _map_excel_columns(headers: list) -> dict:
    """Excel basliklarini standart alanlara esle."""
    aliases = {
        "email": ["email", "e-posta", "eposta", "e_posta", "mail", "e-mail"],
        "name": ["name", "firma", "firma adi", "firma adı", "company", "isim", "ad", "sirket", "şirket"],
        "phone": ["phone", "telefon", "tel", "gsm", "cep"],
        "domain": ["domain", "alan adi", "alan adı"],
        "sector": ["sector", "sektor", "sektör", "kategori", "category"],
        "website": ["website", "web", "site", "url", "web sitesi"],
        "instagram": ["instagram", "ig", "insta"],
        "facebook": ["facebook", "fb"],
        "linkedin": ["linkedin", "li"],
        "date": ["date", "tarih", "kayit tarihi"],
    }
    col_map = {}
    for field, names in aliases.items():
        for i, h in enumerate(headers):
            h_clean = h.strip().lower().replace("İ", "i").replace("ı", "i").replace("ö", "o").replace("ü", "u").replace("ş", "s").replace("ç", "c").replace("ğ", "g")
            for name in names:
                if h_clean == name or h_clean.replace(" ", "_") == name.replace(" ", "_"):
                    col_map[field] = i
                    break
            if field in col_map:
                break
    return col_map


@app.route("/api/database/sources")
def get_sources():
    """Bagli veri kaynaklarini listele."""
    sources = [{"name": "Scanner (Google Maps)", "count": "-", "type": "scanner"}]
    lead_url = app.config.get("LEAD_SHEET_URL", "")
    if lead_url:
        try:
            lead_data = _fetch_lead_sheet(None, lead_url)
            sources.append({"name": "Lead Sheet", "count": len(lead_data), "type": "lead", "url": lead_url})
        except Exception:
            sources.append({"name": "Lead Sheet (hata)", "count": 0, "type": "lead"})
    return jsonify({"sources": sources})


@app.route("/api/database")
def get_database():
    """Tum veri havuzunu getir: Scanner + Lead kaynakları."""
    try:
        sheets = SheetsManager()
        all_data = []

        # 1. Scanner verilerini al
        scanner_rows = sheets.worksheet.get_all_values()
        # Sutun sirasi: Tarih, Sektor, Firma Adi, Telefon, E-posta, Domain, Web Sitesi, Instagram, Facebook, LinkedIn
        for i, row in enumerate(scanner_rows[1:], 2):
            if not row or not any(row):
                continue
            all_data.append({
                "id": f"scanner-{i}",
                "source": "Scanner",
                "date": row[0] if len(row) > 0 else "",
                "sector": row[1] if len(row) > 1 else "",
                "name": row[2] if len(row) > 2 else "",
                "phone": row[3] if len(row) > 3 else "",
                "email": row[4] if len(row) > 4 else "",
                "domain": row[5] if len(row) > 5 else "",
                "website": row[6] if len(row) > 6 else "",
                "instagram": row[7] if len(row) > 7 else "",
                "facebook": row[8] if len(row) > 8 else "",
                "linkedin": row[9] if len(row) > 9 else "",
                "sheet_row": i,
                "sheet_name": "Scanner",
            })

        # 2. Lead Sheet varsa verileri al
        lead_url = app.config.get("LEAD_SHEET_URL", "")
        if lead_url:
            try:
                lead_data = _fetch_lead_sheet(sheets, lead_url)
                all_data.extend(lead_data)
            except Exception as e:
                logger.warning(f"Lead sheet okunamadi: {e}")

        return jsonify({"data": all_data, "total": len(all_data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/database/columns")
def database_columns():
    """Scanner tablosunun gercek baslik satirini dondur."""
    try:
        from config import SHEET_COLUMNS
        sheets = SheetsManager()
        first_row = sheets.worksheet.row_values(1)
        columns = [c.strip() for c in first_row if c.strip()] if first_row else []
        if not columns:
            columns = list(SHEET_COLUMNS)
        return jsonify({"columns": columns})
    except Exception as e:
        logger.error(f"Database columns hatasi: {e}")
        try:
            from config import SHEET_COLUMNS
            return jsonify({"columns": list(SHEET_COLUMNS)})
        except Exception:
            return jsonify({"columns": ["Tarih", "Sektör", "Firma Adı", "Telefon", "E-posta", "Domain", "Web Sitesi", "Instagram", "Facebook", "LinkedIn"]})


def _fetch_lead_sheet(sheets_mgr, lead_url: str, sheet_name: str = None) -> list:
    """Harici lead sheet'inden SADECE secilen sheet'in verilerini cek.

    Diger sheet'lere kesinlikle dokunulmaz.
    """
    import gspread
    from config import get_google_credentials

    # Sheet adi belirtilmemisse config'den al
    if not sheet_name:
        sheet_name = app.config.get("LEAD_SHEET_NAME", "")

    creds = get_google_credentials(["https://www.googleapis.com/auth/spreadsheets.readonly", "https://www.googleapis.com/auth/drive.readonly"])
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_url(lead_url)

    data = []

    # SADECE secilen sheet'i oku
    try:
        if sheet_name:
            ws = spreadsheet.worksheet(sheet_name)
        else:
            # Sheet adi yoksa ilk sheet'i al
            ws = spreadsheet.sheet1
    except Exception as e:
        logger.error(f"Sheet bulunamadi: {sheet_name} - {e}")
        return data

    rows = ws.get_all_values()
    if not rows:
        return data

    headers = [h.lower().strip() for h in rows[0]]
    col_map = _map_columns(headers)

    for i, row in enumerate(rows[1:], 2):
        if not row or not any(row):
            continue

        entry = {
            "id": f"lead-{ws.title}-{i}",
            "source": f"Lead ({ws.title})",
            "date": _get_col(row, col_map, "date"),
            "sector": _get_col(row, col_map, "sector"),
            "name": _get_col(row, col_map, "name"),
            "phone": _get_col(row, col_map, "phone"),
            "email": _get_col(row, col_map, "email"),
            "domain": _get_col(row, col_map, "domain"),
            "website": _get_col(row, col_map, "website"),
            "instagram": _get_col(row, col_map, "instagram"),
            "facebook": _get_col(row, col_map, "facebook"),
            "linkedin": _get_col(row, col_map, "linkedin"),
            "sheet_row": i,
            "sheet_name": ws.title,
        }
        if entry["name"] or entry["email"] or entry["phone"]:
            data.append(entry)

    return data


def _map_columns(headers: list) -> dict:
    """Farkli baslik isimlerini standart alanlara esle."""
    mapping = {
        "date": ["tarih", "date", "kayıt tarihi", "kayit tarihi", "created", "timestamp", "zaman"],
        "sector": ["sektör", "sektor", "sector", "kategori", "industry", "sektoru", "sektörü"],
        "name": ["firma adı", "firma adi", "şirket adı", "sirket adi", "şirket", "sirket",
                 "firma", "company", "ad", "isim", "name", "müşteri", "musteri",
                 "işletme", "isletme", "marka", "unvan", "ünvan"],
        "phone": ["telefon", "telefon numarası", "telefon numarasi", "tel", "phone",
                  "gsm", "cep", "numara", "no", "tel no", "cep telefonu", "iletişim no"],
        "email": ["e-posta", "eposta", "e-mail", "email", "mail", "e-mailadresi",
                  "e-posta adresi", "mailadresi", "elektronik posta"],
        "domain": ["domain", "alan adı", "alan adi"],
        "website": ["web sitesi", "website", "site", "url", "web", "web adresi", "internet sitesi"],
        "instagram": ["instagram", "ig", "insta"],
        "facebook": ["facebook", "fb"],
        "linkedin": ["linkedin", "li"],
    }

    # Turkce karakterleri normalize et
    def normalize(s):
        return (s.lower().strip()
                .replace("İ", "i").replace("ı", "i")
                .replace("Ş", "s").replace("ş", "s")
                .replace("Ç", "c").replace("ç", "c")
                .replace("Ğ", "g").replace("ğ", "g")
                .replace("Ö", "o").replace("ö", "o")
                .replace("Ü", "u").replace("ü", "u"))

    col_map = {}
    norm_headers = [normalize(h) for h in headers]

    for field, aliases in mapping.items():
        if field in col_map:
            continue
        norm_aliases = [normalize(a) for a in aliases]
        for i, nh in enumerate(norm_headers):
            # Tam eslesme
            if nh in norm_aliases:
                col_map[field] = i
                break
            # Kismi eslesme (baslik alias'i iceriyorsa)
            for na in norm_aliases:
                if na in nh or nh in na:
                    col_map[field] = i
                    break
            if field in col_map:
                break

    logger.info(f"Sutun eslestirme: {headers} -> {col_map}")
    return col_map


def _get_col(row: list, col_map: dict, field: str) -> str:
    """Satirdan belirli bir alani al."""
    idx = col_map.get(field)
    if idx is not None and idx < len(row):
        return row[idx].strip()
    return ""


@app.route("/api/database/list-sheets", methods=["POST"])
def list_sheets():
    """URL'deki tablonun sheet'lerini listele."""
    data = request.json
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL boş olamaz"}), 400

    try:
        import gspread
        from config import get_google_credentials

        creds = get_google_credentials(["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"])
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url(url)

        sheets_list = []
        for ws in spreadsheet.worksheets():
            rows = ws.get_all_values()
            row_count = len(rows) - 1 if len(rows) > 0 else 0
            headers = rows[0] if rows else []
            sheets_list.append({
                "title": ws.title,
                "rows": max(row_count, 0),
                "headers": headers[:10],  # Ilk 10 baslik
            })

        return jsonify({
            "spreadsheet_title": spreadsheet.title,
            "sheets": sheets_list,
            "url": url,
        })
    except Exception as e:
        return jsonify({"error": f"Bağlantı hatası: {str(e)}"}), 400


@app.route("/api/database/connect-lead", methods=["POST"])
def connect_lead_sheet():
    """Secilen sheet'i bagla."""
    data = request.json
    url = data.get("url", "").strip()
    sheet_name = data.get("sheet_name", "").strip()
    if not url:
        return jsonify({"error": "URL boş olamaz"}), 400
    if not sheet_name:
        return jsonify({"error": "Sheet seçilmedi"}), 400

    # URL ve sheet adini kaydet
    app.config["LEAD_SHEET_URL"] = url
    app.config["LEAD_SHEET_NAME"] = sheet_name

    # .env dosyasina kaydet
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        if os.path.exists(env_path):
            with open(env_path, "r") as f:
                lines = f.readlines()
            lines = [l for l in lines if not l.startswith("LEAD_SHEET_URL=") and not l.startswith("LEAD_SHEET_NAME=")]
            with open(env_path, "w") as f:
                f.writelines(lines)

        with open(env_path, "a") as f:
            f.write(f"LEAD_SHEET_URL={url}\n")
            f.write(f"LEAD_SHEET_NAME={sheet_name}\n")
    except Exception:
        pass

    # Test - sadece secilen sheet'i oku
    try:
        lead_data = _fetch_lead_sheet(None, url, sheet_name)
        return jsonify({
            "message": f"'{sheet_name}' bağlandı! {len(lead_data)} kayıt bulundu.",
            "count": len(lead_data),
        })
    except Exception as e:
        return jsonify({"error": f"Bağlantı hatası: {str(e)}"}), 400


@app.route("/api/import/parse-file", methods=["POST"])
def import_parse_file():
    """Yuklenen dosyayi parse et, baslik + ilk 200 satiri dondur."""
    if "file" not in request.files:
        return jsonify({"error": "Dosya secilmedi."}), 400
    file = request.files["file"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        return jsonify({"error": "Desteklenen: .xlsx, .xls, .csv"}), 400
    try:
        import io
        if ext == "csv":
            import csv
            content = file.read().decode("utf-8-sig")
            reader = csv.reader(io.StringIO(content))
            all_rows = list(reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            all_rows = [[str(c or "").strip() for c in row] for row in ws.iter_rows(values_only=True)]

        if not all_rows:
            return jsonify({"error": "Dosya bos."}), 400

        headers = [str(h or "").strip() for h in all_rows[0]]
        data_rows = []
        for row in all_rows[1:201]:
            row_list = [str(c or "").strip() for c in row]
            if any(row_list):
                data_rows.append(row_list)

        return jsonify({
            "headers": headers,
            "rows": data_rows,
            "total": len(all_rows) - 1,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/import/fetch-sheet", methods=["POST"])
def import_fetch_sheet():
    """Google Sheet'teki belirli sheet'in verilerini dondur."""
    data = request.json
    url = data.get("url", "").strip()
    sheet_name = data.get("sheet_name", "").strip()
    if not url or not sheet_name:
        return jsonify({"error": "URL ve sheet_name gerekli"}), 400
    try:
        import gspread
        from config import get_google_credentials
        creds = get_google_credentials([
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ])
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url(url)
        ws = spreadsheet.worksheet(sheet_name)
        all_values = ws.get_all_values()
        if not all_values:
            return jsonify({"error": "Sheet bos."}), 400
        headers = all_values[0]
        rows = all_values[1:201]
        return jsonify({"headers": headers, "rows": rows, "total": len(all_values) - 1})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/import/execute", methods=["POST"])
def import_execute():
    """Import edilen satirlari Leads sheet'e yaz."""
    data = request.json
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"error": "Veri yok"}), 400
    try:
        from datetime import datetime as _dt
        today = _dt.now().strftime("%Y-%m-%d")
        leads_mgr = LeadsManager()
        leads = []
        seen_emails = set()

        for r in rows:
            first = r.get("Adı", "").strip()
            last  = r.get("Soyadı", "").strip()
            email = (r.get("Email Adresi", "") or r.get("E-posta", "")).strip()
            phone = (r.get("Telefon Numarası", "") or r.get("Telefon", "")).strip()
            city  = r.get("Şehir", "").strip()

            if not first and not last and not email and not phone:
                continue

            if email:
                key = email.lower()
                if key in seen_emails:
                    continue
                seen_emails.add(key)

            leads.append({
                "date":       r.get("Tarih", today) or today,
                "first_name": first,
                "last_name":  last,
                "phone":      phone,
                "email":      email,
                "city":       city,
            })

        total_valid = len(leads)
        added = leads_mgr.append_leads(leads)

        # Auto-trigger: import kaynaklı lead_created kampanyalarını tetikle
        if added > 0:
            try:
                from marketing.campaigns import auto_trigger_leads
                trigger_leads = [
                    {
                        "email":      l.get("email", ""),
                        "name":       f"{l.get('first_name','')} {l.get('last_name','')}".strip() or l.get("email",""),
                        "first_name": l.get("first_name", ""),
                        "last_name":  l.get("last_name", ""),
                        "city":       l.get("city", ""),
                        "sector":     "",
                        "domain":     "",
                        "website":    "",
                    }
                    for l in leads if l.get("email")
                ]
                n = auto_trigger_leads(trigger_leads, "import")
                if n:
                    logger.info(f"auto_trigger: {n} import lead için kampanya tetiklendi")
            except Exception as te:
                logger.debug(f"auto_trigger_import error: {te}")

        return jsonify({
            "created": added,
            "skipped": total_valid - added,
            "total":   len(rows),
        })
    except Exception as e:
        logger.error(f"Import execute hatasi: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads")
def get_leads():
    """Leads sheet'indeki tum kayitlari dondur."""
    try:
        leads_mgr = LeadsManager()
        data = leads_mgr.get_all_leads()
        return jsonify({"data": data, "total": len(data)})
    except Exception as e:
        logger.error(f"Leads getirme hatasi: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads/delete", methods=["POST"])
def delete_leads():
    """Leads sheet'inden satir sil."""
    data = request.json
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"error": "Silinecek satir secilmedi"}), 400
    try:
        leads_mgr = LeadsManager()
        deleted = leads_mgr.delete_rows([int(r) for r in rows])
        return jsonify({"deleted": deleted})
    except Exception as e:
        logger.error(f"Lead silme hatasi: {e}")
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  GOOGLE OAUTH (kullanici kendi hesabiyla baglanir)
# ─────────────────────────────────────────────────────────────

GOOGLE_CLIENT_ID     = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
GOOGLE_API_KEY       = os.getenv("GOOGLE_API_KEY", "")   # public API key for Picker widget
GOOGLE_APP_ID        = os.getenv("GOOGLE_APP_ID", "")    # GCP project number for Picker app context
GOOGLE_SCOPES = "https://www.googleapis.com/auth/drive.file"
GOOGLE_TOKEN_COOKIE = "google_sheets_token"


def _get_google_token():
    return request.cookies.get(GOOGLE_TOKEN_COOKIE, "")


def _redirect_uri():
    base = os.getenv("APP_BASE_URL", "").rstrip("/") or request.host_url.rstrip("/")
    return f"{base}/api/google/callback"


@app.route("/api/google/status")
def google_status():
    token = _get_google_token()
    return jsonify({"connected": bool(token)})


@app.route("/api/google/connect")
def google_connect():
    import secrets, urllib.parse
    if not GOOGLE_CLIENT_ID:
        return jsonify({"error": "GOOGLE_CLIENT_ID ayarlanmamis"}), 500
    state = secrets.token_urlsafe(32)
    session["google_oauth_state"] = state
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": _redirect_uri(),
        "response_type": "code",
        "scope": GOOGLE_SCOPES,
        "prompt": "consent select_account",
        "state": state,
    }
    url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
    return redirect(url)


@app.route("/api/google/callback")
def google_callback():
    import urllib.parse, urllib.request as ureq
    error = request.args.get("error")
    code  = request.args.get("code")
    state = request.args.get("state")
    back  = "/external"

    if error or not code:
        return redirect(back + "?google_error=" + urllib.parse.quote(error or "cancelled"))

    if state != session.pop("google_oauth_state", None):
        return redirect(back + "?google_error=invalid_state")

    # Token exchange
    payload = urllib.parse.urlencode({
        "code": code,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "redirect_uri": _redirect_uri(),
        "grant_type": "authorization_code",
    }).encode()
    try:
        req = ureq.Request("https://oauth2.googleapis.com/token", data=payload,
                           headers={"Content-Type": "application/x-www-form-urlencoded"})
        with ureq.urlopen(req) as resp:
            tokens = json.loads(resp.read())
    except Exception as e:
        return redirect(back + "?google_error=token_exchange_failed")

    access_token = tokens.get("access_token", "")
    expires_in   = int(tokens.get("expires_in", 3600))

    if not access_token:
        return redirect(back + "?google_error=no_access_token")

    resp = make_response(redirect(back + "?google=connected"))
    resp.set_cookie(GOOGLE_TOKEN_COOKIE, access_token,
                    httponly=True, samesite="Lax",
                    max_age=expires_in, secure=request.is_secure)
    return resp


@app.route("/api/google/disconnect", methods=["POST"])
def google_disconnect():
    resp = make_response(jsonify({"success": True}))
    resp.set_cookie(GOOGLE_TOKEN_COOKIE, "", max_age=0)
    return resp


@app.route("/api/google/picker-config")
def google_picker_config():
    """Returns non-secret Picker configuration for the frontend widget."""
    return jsonify({
        "api_key": GOOGLE_API_KEY,
        "app_id":  GOOGLE_APP_ID,
    })


@app.route("/api/google/picker-token")
def google_picker_token():
    """Exposes the OAuth access token to the frontend for use in the Google Picker widget.
    Required because the token is stored in an httpOnly cookie."""
    token = _get_google_token()
    if not token:
        return jsonify({"error": "google_not_connected"}), 401
    return jsonify({"access_token": token})


@app.route("/api/google/sheets")
def google_list_sheets():
    token = _get_google_token()
    if not token:
        return jsonify({"error": "google_not_connected"}), 401
    q_param = request.args.get("q", "").strip()
    import urllib.request as ureq, urllib.parse
    query_parts = [
        "mimeType='application/vnd.google-apps.spreadsheet'",
        "trashed=false",
    ]
    if q_param:
        safe = q_param.replace("'", "\\'")
        query_parts.append(f"name contains '{safe}'")
    params = urllib.parse.urlencode({
        "q": " and ".join(query_parts),
        "fields": "files(id,name,modifiedTime)",
        "orderBy": "modifiedTime desc",
        "pageSize": "50",
    })
    url = "https://www.googleapis.com/drive/v3/files?" + params
    req = ureq.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with ureq.urlopen(req) as resp:
            data = json.loads(resp.read())
        return jsonify({"files": data.get("files", [])})
    except Exception as e:
        code = getattr(e, "code", 500)
        if code == 401:
            return jsonify({"error": "google_token_expired"}), 401
        return jsonify({"error": str(e)}), 500


@app.route("/api/google/sheets/<spreadsheet_id>")
def google_sheet_data(spreadsheet_id):
    token = _get_google_token()
    if not token:
        return jsonify({"error": "google_not_connected"}), 401
    sheet_name = request.args.get("sheet", "").strip()
    import urllib.request as ureq, urllib.parse
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
    headers_auth = {"Authorization": f"Bearer {token}"}

    try:
        if not sheet_name:
            # Sekmeleri listele
            req = ureq.Request(f"{base}?fields=sheets.properties(sheetId,title,index)",
                               headers=headers_auth)
            with ureq.urlopen(req) as resp:
                data = json.loads(resp.read())
            tabs = [{"id": s["properties"]["sheetId"],
                     "title": s["properties"]["title"],
                     "index": s["properties"]["index"]}
                    for s in data.get("sheets", [])]
            return jsonify({"tabs": tabs})
        else:
            # Satir verilerini getir
            range_enc = urllib.parse.quote(sheet_name)
            req = ureq.Request(f"{base}/values/{range_enc}", headers=headers_auth)
            with ureq.urlopen(req) as resp:
                data = json.loads(resp.read())
            values = data.get("values", [])
            if not values:
                return jsonify({"headers": [], "rows": [], "total": 0})
            hdrs = [str(h).strip() for h in values[0]]
            rows = [[str(c) for c in row] for row in values[1:201]]
            return jsonify({"headers": hdrs, "rows": rows, "total": len(values) - 1})
    except Exception as e:
        code = getattr(e, "code", 500)
        if code == 401:
            return jsonify({"error": "google_token_expired"}), 401
        return jsonify({"error": str(e)}), 500


# Baslangicta lead ayarlarini .env'den yukle
import os
_lead_url = os.getenv("LEAD_SHEET_URL", "")
_lead_sheet = os.getenv("LEAD_SHEET_NAME", "")
if _lead_url:
    app.config["LEAD_SHEET_URL"] = _lead_url
if _lead_sheet:
    app.config["LEAD_SHEET_NAME"] = _lead_sheet


# --- Resend Webhook ---
@app.route("/api/webhooks/resend", methods=["POST"])
def resend_webhook():
    """Resend delivery events → email_events tablosu + otomatik suppression."""
    # Svix imza doğrulaması (Resend, Svix kullanır)
    secret = os.getenv("RESEND_WEBHOOK_SECRET", "")
    if secret:
        import hmac, hashlib, base64
        svix_id        = request.headers.get("svix-id", "")
        svix_timestamp = request.headers.get("svix-timestamp", "")
        svix_signature = request.headers.get("svix-signature", "")
        if not (svix_id and svix_timestamp and svix_signature):
            return jsonify({"error": "missing svix headers"}), 401
        signed_content = f"{svix_id}.{svix_timestamp}.{request.get_data(as_text=True)}"
        key = base64.b64decode(secret.removeprefix("whsec_"))
        expected = "v1," + base64.b64encode(
            hmac.new(key, signed_content.encode(), hashlib.sha256).digest()
        ).decode()
        if not any(sig == expected for sig in svix_signature.split(" ")):
            return jsonify({"error": "invalid signature"}), 401

    payload = request.json or {}
    event_type = payload.get("type", "")
    data = payload.get("data", {})
    to_email = ""

    # Resend webhook payload yapısı: data.to = [email] veya data.email
    if isinstance(data.get("to"), list) and data["to"]:
        to_email = data["to"][0].lower().strip()
    elif data.get("email_id"):
        to_email = data.get("to", "")

    if not to_email:
        return jsonify({"ok": True})

    provider_message_id = data.get("email_id", "")

    from marketing.db import sb_select, sb_insert
    queue_id = None
    campaign_id = None
    try:
        rows = sb_select("email_queue", {
            "provider_message_id": f"eq.{provider_message_id}",
            "select": "id,campaign_id",
        })
        if rows:
            queue_id   = rows[0]["id"]
            campaign_id = rows[0].get("campaign_id")
    except Exception:
        pass

    # Event kaydet
    try:
        sb_insert("email_events", {
            "queue_id":    queue_id,
            "campaign_id": campaign_id,
            "to_email":    to_email,
            "event_type":  event_type,
            "metadata":    data,
        })
    except Exception as e:
        logger.warning(f"resend_webhook event insert failed: {e}")

    # Bounce / complained → otomatik suppression
    if event_type in ("email.bounced", "email.complained"):
        reason = "bounced" if event_type == "email.bounced" else "complained"
        add_suppression(to_email, reason=reason, source="webhook")

    return jsonify({"ok": True})


start_scheduler()

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5050))
    print("\n" + "=" * 50)
    print("  POWER SCANNER")
    print(f"  Dashboard: http://localhost:{port}")
    print("=" * 50 + "\n")
    socketio.run(app, debug=False, port=port, host="0.0.0.0", allow_unsafe_werkzeug=True)
